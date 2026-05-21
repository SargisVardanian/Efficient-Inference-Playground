from __future__ import annotations

import argparse
import subprocess
import tempfile
from pathlib import Path

import pandas as pd

import sys

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from eip.io_utils import ensure_parent


SARGIS_HARDWARE = "Apple M3 Pro"
SILVI_HARDWARE = "Apple M4 Max"
SILVI_BRANCH = "origin/experiments/baseline-kvcache"
SILVI_XLSX_PATH = "results/benchmark_results.xlsx"


SARGIS_RAW_INPUTS = [
    "results/raw/hf_baseline_all_clean.csv",
    "results/raw/ollama_quantized_short.csv",
    "results/raw/ollama_quantized_medium.csv",
    "results/raw/ollama_quantized_long.csv",
    "results/raw/hf_kv_short.csv",
    "results/raw/hf_kv_medium.csv",
    "results/raw/hf_kv_long.csv",
]


def load_sargis_all_runs() -> pd.DataFrame:
    frames = []
    for item in SARGIS_RAW_INPUTS:
        path = Path(item)
        if not path.exists() or path.stat().st_size == 0:
            continue
        df = pd.read_csv(path)
        if len(df):
            frames.append(df)
    if not frames:
        raise SystemExit("No Sargis result CSV files found.")

    df = pd.concat(frames, ignore_index=True)
    df = df[df["experiment"].notna() & df["prompt_id"].notna()].copy()
    df = df.drop_duplicates(["experiment", "prompt_id", "run"], keep="last")
    df = df.rename(columns={"prompt_tokens": "input_tokens"})
    for col in [
        "technique",
        "backend",
        "model",
        "length_bucket",
        "task",
        "dataset",
        "run",
        "input_tokens",
        "generated_tokens",
        "latency_s",
        "tokens_per_second",
        "rss_before_mb",
        "rss_after_mb",
        "output",
        "reference",
        "quantization_method",
        "cache_policy",
        "device_runtime_note",
    ]:
        if col not in df.columns:
            df[col] = None
    df["owner"] = "Sargis"
    df["hardware"] = SARGIS_HARDWARE
    df["benchmark_suite"] = "Needle-in-a-haystack retrieval"
    return df


def normalize(text: object) -> str:
    return " ".join(str(text or "").strip().lower().split())


def build_sargis_quality(all_runs: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for row in all_runs.to_dict("records"):
        reference = normalize(row.get("reference"))
        output = normalize(row.get("output"))
        has_reference = bool(reference)
        rows.append(
            {
                "owner": "Sargis",
                "hardware": SARGIS_HARDWARE,
                "experiment": row["experiment"],
                "technique": row["technique"],
                "backend": row["backend"],
                "model": row.get("model"),
                "prompt_id": row["prompt_id"],
                "length_bucket": row["length_bucket"],
                "task": row.get("task"),
                "run": row["run"],
                "reference_exact_match": float(output == reference) if has_reference else None,
                "reference_contains_match": float(reference in output) if has_reference else None,
                "rougeL_to_baseline": None,
                "char_similarity_to_baseline": None,
                "exact_match_to_baseline": None,
            }
        )
    return pd.DataFrame(rows)


def infer_sargis_note(row: pd.Series | pd.DataFrame) -> str:
    if isinstance(row, pd.DataFrame):
        row = row.iloc[0]
    existing = str(row.get("device_runtime_note", "") or "").strip()
    backend = str(row.get("backend", "") or "")
    cache_policy = str(row.get("cache_policy", "") or "")
    quantization = str(row.get("quantization_method", "") or "")
    length = str(row.get("length_bucket", "") or "")
    if cache_policy == "recency_window" and existing.lower().startswith("transformers default cache"):
        existing = ""
    if existing and existing.lower() != "nan":
        return existing
    if backend == "ollama" or quantization == "gguf_q4_k_m":
        return "Mixed-backend Ollama GGUF Q4_K_M fallback; not directly comparable to HF device-path rows."
    if cache_policy == "recency_window" and length == "long":
        return "HF Transformers hybrid recency-window KV; long prompts used CPU fallback policy."
    if cache_policy == "recency_window":
        return "HF Transformers hybrid recency-window KV on Apple Silicon MPS."
    if backend == "transformers" and length == "long":
        return "HF Transformers baseline; long prompts used CPU fallback policy."
    if backend == "transformers":
        return "HF Transformers baseline on Apple Silicon MPS."
    return ""


def build_summary(all_runs: pd.DataFrame, quality: pd.DataFrame) -> pd.DataFrame:
    df = all_runs.copy()
    for col in ["latency_s", "tokens_per_second", "input_tokens", "generated_tokens", "rss_before_mb", "rss_after_mb"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df["rss_delta_mb"] = df["rss_after_mb"] - df["rss_before_mb"]

    q = quality.copy()
    metric_cols = [
        "reference_exact_match",
        "reference_contains_match",
        "rougeL_to_baseline",
        "char_similarity_to_baseline",
        "exact_match_to_baseline",
    ]
    for col in metric_cols:
        if col in q.columns:
            q[col] = pd.to_numeric(q[col], errors="coerce")

    group_cols = ["owner", "hardware", "experiment", "technique", "backend", "length_bucket", "task"]
    perf = (
        df.groupby(group_cols, dropna=False)
        .agg(
            latency_p50=("latency_s", "median"),
            latency_p99=("latency_s", lambda values: values.quantile(0.99)),
            tokens_per_second=("tokens_per_second", "mean"),
            generated_tokens=("generated_tokens", "mean"),
            input_tokens=("input_tokens", "mean"),
            rss_delta_peak_mb=("rss_delta_mb", "max"),
            rss_after_peak_mb=("rss_after_mb", "max"),
            n_runs=("latency_s", "count"),
        )
        .reset_index()
    )
    qual = (
        q.groupby(group_cols, dropna=False)[metric_cols]
        .mean(numeric_only=True)
        .reset_index()
    )
    out = perf.merge(qual, on=group_cols, how="left")
    if "device_runtime_note" in df.columns:
        notes = df.drop_duplicates(group_cols).copy()
        notes["runtime_note"] = notes.apply(infer_sargis_note, axis=1)
        notes = notes[group_cols + ["runtime_note"]]
        out = out.merge(notes, on=group_cols, how="left")
    return out.round(4)


def read_silvi_workbook_from_git() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    with tempfile.TemporaryDirectory() as td:
        target = Path(td) / "silvi_benchmark_results.xlsx"
        data = subprocess.check_output(["git", "show", f"{SILVI_BRANCH}:{SILVI_XLSX_PATH}"])
        target.write_bytes(data)
        summary = pd.read_excel(target, sheet_name="Summary")
        quality = pd.read_excel(target, sheet_name="Quality (per run)")
        all_runs = pd.read_excel(target, sheet_name="All Runs")
    for df in [summary, quality, all_runs]:
        df.insert(0, "hardware", SILVI_HARDWARE)
        df.insert(0, "owner", "Silvi")
    summary["benchmark_suite"] = "CNN/GovReport summarization + GSM8K reasoning"
    all_runs["benchmark_suite"] = "CNN/GovReport summarization + GSM8K reasoning"
    summary["runtime_note"] = summary.apply(
        lambda row: "Ollama Gemma4 Q4_K_M on M4 Max."
        if row.get("backend") == "ollama"
        else "HF Qwen SinkCache/sliding-window comparison on M4 Max.",
        axis=1,
    )
    return summary, quality, all_runs


def make_pivots(summary: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    index = ["owner", "hardware", "experiment"]
    latency = summary.pivot_table(index=index, columns="length_bucket", values="latency_p50", aggfunc="median").reset_index()
    throughput = summary.pivot_table(index=index, columns="length_bucket", values="tokens_per_second", aggfunc="median").reset_index()
    quality_col = "reference_contains_match" if "reference_contains_match" in summary.columns else "rougeL_to_baseline"
    quality = summary.pivot_table(index=index, columns="length_bucket", values=quality_col, aggfunc="mean").reset_index()
    return latency, throughput, quality


def style_workbook(writer: pd.ExcelWriter) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_bg = "1F4E79"
    header_fg = "FFFFFF"
    for ws in writer.sheets.values():
        for cell in ws[1]:
            cell.font = Font(bold=True, color=header_fg)
            cell.fill = PatternFill("solid", fgColor=header_bg)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        for col in ws.columns:
            width = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 4, 42)
        ws.freeze_panes = "A2"


def write_workbook(path: Path, summary: pd.DataFrame, quality: pd.DataFrame, all_runs: pd.DataFrame, notes: pd.DataFrame) -> None:
    latency, throughput, quality_pivot = make_pivots(summary)
    ensure_parent(path)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        latency.to_excel(writer, sheet_name="Latency p50 (s)", index=False)
        throughput.to_excel(writer, sheet_name="Throughput (tok-s)", index=False)
        quality_pivot.to_excel(writer, sheet_name="Quality pivot", index=False)
        quality.to_excel(writer, sheet_name="Quality (per run)", index=False)
        all_runs.drop(columns=["output"], errors="ignore").to_excel(writer, sheet_name="All Runs", index=False)
        notes.to_excel(writer, sheet_name="Notes", index=False)
        style_workbook(writer)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--sargis-xlsx", default="results/sargis_needle_benchmark_results.xlsx")
    parser.add_argument("--team-xlsx", default="results/team_final_benchmark_results.xlsx")
    parser.add_argument("--sargis-summary-csv", default="results/processed/sargis_needle_final_summary.csv")
    parser.add_argument("--team-summary-csv", default="results/processed/team_final_summary.csv")
    args = parser.parse_args()

    sargis_all = load_sargis_all_runs()
    sargis_quality = build_sargis_quality(sargis_all)
    sargis_summary = build_summary(sargis_all, sargis_quality)
    sargis_notes = pd.DataFrame(
        [
            {
                "owner": "Sargis",
                "hardware": SARGIS_HARDWARE,
                "note": "Baseline/KV are HF Transformers; quantized row is mixed-backend Ollama GGUF Q4_K_M fallback.",
            },
            {
                "owner": "Sargis",
                "hardware": SARGIS_HARDWARE,
                "note": "Needle has long input and very short output, so KV recency-window mainly exposes prefill/decode workload mismatch.",
            },
            {
                "owner": "Sargis",
                "hardware": SARGIS_HARDWARE,
                "note": "RSS and system memory are process/system proxies, not true model VRAM.",
            },
        ]
    )

    ensure_parent(args.sargis_summary_csv)
    sargis_summary.to_csv(args.sargis_summary_csv, index=False)
    sargis_all.to_csv("results/processed/sargis_needle_all_runs.csv", index=False)
    sargis_quality.to_csv("results/processed/sargis_needle_final_quality.csv", index=False)
    write_workbook(Path(args.sargis_xlsx), sargis_summary, sargis_quality, sargis_all, sargis_notes)

    silvi_summary, silvi_quality, silvi_all = read_silvi_workbook_from_git()
    team_summary = pd.concat([sargis_summary, silvi_summary], ignore_index=True, sort=False)
    team_quality = pd.concat([sargis_quality, silvi_quality], ignore_index=True, sort=False)
    team_all = pd.concat([sargis_all, silvi_all], ignore_index=True, sort=False)
    team_notes = pd.concat(
        [
            sargis_notes,
            pd.DataFrame(
                [
                    {
                        "owner": "Silvi",
                        "hardware": SILVI_HARDWARE,
                        "note": "Silvi workbook sourced from origin/experiments/baseline-kvcache:results/benchmark_results.xlsx.",
                    },
                    {
                        "owner": "Silvi",
                        "hardware": SILVI_HARDWARE,
                        "note": "Silvi main Ollama baseline/KV rows use Gemma4 Q4_K_M on M4 Max; HF SinkCache rows use Qwen on M4 Max.",
                    },
                ]
            ),
        ],
        ignore_index=True,
    )

    ensure_parent(args.team_summary_csv)
    team_summary.to_csv(args.team_summary_csv, index=False)
    team_all.to_csv("results/processed/team_final_all_runs.csv", index=False)
    team_quality.to_csv("results/processed/team_final_quality.csv", index=False)
    write_workbook(Path(args.team_xlsx), team_summary, team_quality, team_all, team_notes)

    print(f"Wrote {args.sargis_xlsx}")
    print(f"Wrote {args.team_xlsx}")
    print(f"Wrote {args.sargis_summary_csv}")
    print(f"Wrote {args.team_summary_csv}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
