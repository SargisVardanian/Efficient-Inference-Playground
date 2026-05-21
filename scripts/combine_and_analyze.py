"""
Combine all three experiment CSV files into one master analysis spreadsheet
with quality metrics, latency percentiles, throughput, and a styled Excel report.

Input files (expected under results/raw/):
  - ollama_benchmark.csv     (baseline_gemma4_e4b)
  - kvcache_benchmark.csv    (kvcache_limited_512)
  - hf_sinkcache.csv         (hf_baseline, hf_sinkcache)

Output:
  - results/processed/all_runs.csv             (every row, harmonised schema)
  - results/processed/quality_metrics.csv      (ROUGE-L, char sim, exact match vs same-family baseline)
  - results/processed/summary.csv              (per (experiment, length_bucket, task) aggregate)
  - results/benchmark_results.xlsx             (styled multi-sheet workbook)
"""
from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
from rouge_score import rouge_scorer

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))
from eip.metrics import char_similarity, exact_match  # noqa: E402

RAW_DIR       = Path("results/raw")
PROCESSED_DIR = Path("results/processed")
XLSX_PATH     = Path("results/benchmark_results.xlsx")


HARMONISED_FIELDS = [
    "experiment", "technique", "backend", "model",
    "prompt_id", "length_bucket", "task",
    "run", "input_tokens", "generated_tokens",
    "latency_s", "tokens_per_second",
    "rss_before_mb", "rss_after_mb",
    "output",
]


def load_ollama(csv_path: Path, backend_label: str) -> pd.DataFrame:
    if not csv_path.exists():
        print(f"  Missing: {csv_path} (skipping)")
        return pd.DataFrame(columns=HARMONISED_FIELDS)
    df = pd.read_csv(csv_path)
    df = df.rename(columns={"prompt_tokens": "input_tokens"})
    df["backend"] = backend_label
    for c in HARMONISED_FIELDS:
        if c not in df.columns:
            df[c] = None
    return df[HARMONISED_FIELDS]


def load_hf(csv_path: Path, backend_label: str) -> pd.DataFrame:
    if not csv_path.exists():
        print(f"  Missing: {csv_path} (skipping)")
        return pd.DataFrame(columns=HARMONISED_FIELDS)
    df = pd.read_csv(csv_path)
    if "run" not in df.columns:
        df["run"] = 1
    df["backend"] = backend_label
    for c in HARMONISED_FIELDS:
        if c not in df.columns:
            df[c] = None
    return df[HARMONISED_FIELDS]


def compute_quality(df: pd.DataFrame) -> pd.DataFrame:
    """Compare each output to the baseline in the SAME backend family."""
    df = df.copy()
    df["output"] = df["output"].fillna("").astype(str)
    scorer = rouge_scorer.RougeScorer(["rougeL"], use_stemmer=True)

    baseline_by_backend = {
        "ollama":      "baseline_gemma4_e4b",
        "huggingface": "hf_baseline",
    }

    rows = []
    for backend, baseline_exp in baseline_by_backend.items():
        sub = df[df["backend"] == backend]
        baseline_outputs = (
            sub[sub["experiment"] == baseline_exp]
            .sort_values(["prompt_id", "run"])
            .groupby("prompt_id")["output"]
            .first()
            .to_dict()
        )
        for r in sub.to_dict("records"):
            base = baseline_outputs.get(r["prompt_id"], "")
            rouge_l = scorer.score(base, r["output"])["rougeL"].fmeasure if base else None
            rows.append({
                "experiment":   r["experiment"],
                "technique":    r["technique"],
                "backend":      backend,
                "model":        r["model"],
                "prompt_id":    r["prompt_id"],
                "length_bucket": r["length_bucket"],
                "task":         r["task"],
                "run":          r["run"],
                "exact_match_to_baseline":   exact_match(base, r["output"]) if base else None,
                "char_similarity_to_baseline": char_similarity(base, r["output"]) if base else None,
                "rougeL_to_baseline":        rouge_l,
            })
    return pd.DataFrame(rows)


def build_summary(perf: pd.DataFrame, quality: pd.DataFrame) -> pd.DataFrame:
    perf = perf.copy()
    quality = quality.copy()
    for col in ("latency_s", "tokens_per_second", "input_tokens", "generated_tokens"):
        if col in perf.columns:
            perf[col] = pd.to_numeric(perf[col], errors="coerce")
    for col in ("rougeL_to_baseline", "char_similarity_to_baseline", "exact_match_to_baseline"):
        if col in quality.columns:
            quality[col] = pd.to_numeric(quality[col], errors="coerce")

    perf_g = perf.groupby(["experiment", "technique", "backend", "length_bucket", "task"]).agg(
        latency_p50=("latency_s", "median"),
        latency_p99=("latency_s", lambda x: x.quantile(0.99)),
        tokens_per_second=("tokens_per_second", "median"),
        generated_tokens=("generated_tokens", "median"),
        input_tokens=("input_tokens", "median"),
        n_runs=("latency_s", "count"),
    ).round(3).reset_index()

    qual_g = quality.groupby(["experiment", "technique", "backend", "length_bucket", "task"]).agg(
        rougeL_to_baseline=("rougeL_to_baseline", "mean"),
        char_similarity_to_baseline=("char_similarity_to_baseline", "mean"),
        exact_match_to_baseline=("exact_match_to_baseline", "mean"),
    ).round(3).reset_index()

    return perf_g.merge(qual_g, on=["experiment", "technique", "backend", "length_bucket", "task"], how="left")


def write_excel(all_runs: pd.DataFrame, quality: pd.DataFrame, summary: pd.DataFrame) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_BG = "1F4E79"
    HEADER_FG = "FFFFFF"

    def style(ws):
        for cell in ws[1]:
            cell.font = Font(bold=True, color=HEADER_FG)
            cell.fill = PatternFill("solid", fgColor=HEADER_BG)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        for col in ws.columns:
            w = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 36)
        ws.freeze_panes = "A2"

    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(XLSX_PATH, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)

        # Latency pivot
        lat = all_runs.groupby(["experiment", "length_bucket"])["latency_s"].median().round(3).unstack()
        lat.to_excel(writer, sheet_name="Latency p50 (s)")

        # Throughput pivot
        tps = all_runs.groupby(["experiment", "length_bucket"])["tokens_per_second"].median().round(1).unstack()
        tps.to_excel(writer, sheet_name="Throughput (tok-s)")

        # Quality pivot
        if "rougeL_to_baseline" in quality.columns:
            q = quality.copy()
            q["rougeL_to_baseline"] = pd.to_numeric(q["rougeL_to_baseline"], errors="coerce")
            rgl = q.groupby(["experiment", "length_bucket"])["rougeL_to_baseline"].mean().round(3).unstack()
            rgl.to_excel(writer, sheet_name="ROUGE-L vs baseline")

        quality.to_excel(writer, sheet_name="Quality (per run)", index=False)
        all_runs.drop(columns=["output"]).to_excel(writer, sheet_name="All Runs", index=False)

        for s in writer.sheets.values():
            style(s)


def main() -> int:
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)

    print("Loading raw CSVs ...")
    dfs = [
        load_ollama(RAW_DIR / "ollama_benchmark.csv",  backend_label="ollama"),
        load_ollama(RAW_DIR / "kvcache_benchmark.csv", backend_label="ollama"),
        load_hf    (RAW_DIR / "hf_sinkcache.csv",      backend_label="huggingface"),
    ]
    all_runs = pd.concat([d for d in dfs if len(d)], ignore_index=True)
    print(f"  Total rows: {len(all_runs)}")
    print(f"  Experiments: {sorted(all_runs['experiment'].dropna().unique())}")

    all_runs.to_csv(PROCESSED_DIR / "all_runs.csv", index=False)

    print("Computing quality metrics ...")
    quality = compute_quality(all_runs)
    quality.to_csv(PROCESSED_DIR / "quality_metrics.csv", index=False)

    print("Building summary ...")
    summary = build_summary(all_runs, quality)
    summary.to_csv(PROCESSED_DIR / "summary.csv", index=False)

    print(f"Writing {XLSX_PATH} ...")
    write_excel(all_runs, quality, summary)

    print("\nDone.")
    print(f"  {PROCESSED_DIR / 'all_runs.csv'}")
    print(f"  {PROCESSED_DIR / 'quality_metrics.csv'}")
    print(f"  {PROCESSED_DIR / 'summary.csv'}")
    print(f"  {XLSX_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
