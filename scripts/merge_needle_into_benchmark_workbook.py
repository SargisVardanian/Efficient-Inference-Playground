from __future__ import annotations

import argparse
import json
import subprocess
import warnings
from pathlib import Path

import pandas as pd


DEFAULT_INPUT_WORKBOOK = "/Users/sargisvardanyan/Downloads/benchmark_results.xlsx"
DEFAULT_OUTPUT_WORKBOOK = "results/benchmark_results.xlsx"
SOURCE_BRANCH = "origin/experiments/baseline-kvcache"
ORIGINAL_HARDWARE = "Apple M4 Max"
NEEDLE_HARDWARE = "Apple M3 Pro"
ORIGINAL_SUITE = "summarization_reasoning"
NEEDLE_SUITE = "needle_in_a_haystack"

RAW_INPUTS = [
    "results/raw/hf_baseline_all_clean.csv",
    "results/raw/ollama_quantized_short.csv",
    "results/raw/ollama_quantized_medium.csv",
    "results/raw/ollama_quantized_long.csv",
    "results/raw/hf_kv_short.csv",
    "results/raw/hf_kv_medium.csv",
    "results/raw/hf_kv_long.csv",
]


EXPERIMENT_RENAMES = {
    "hf_baseline_gemma4_e4b": "needle_hf_baseline_gemma4_e4b",
    "ollama_quantized_gemma4_e4b_q4km": "needle_ollama_quantized_gemma4_e4b_q4km",
    "hf_kv_window_gemma4_e4b": "needle_hf_kv_window_gemma4_e4b",
}


def normalize(text: object) -> str:
    return " ".join(str(text or "").strip().lower().split())


def load_needle_runs() -> pd.DataFrame:
    frames = []
    for item in RAW_INPUTS:
        path = Path(item)
        if path.exists() and path.stat().st_size > 0:
            frames.append(pd.read_csv(path))
    if not frames:
        raise SystemExit("No Needle raw CSV files found.")

    df = pd.concat(frames, ignore_index=True)
    df = df[df["experiment"].notna() & df["prompt_id"].notna()].copy()
    df = df.drop_duplicates(["experiment", "prompt_id", "run"], keep="last")
    df["experiment"] = df["experiment"].replace(EXPERIMENT_RENAMES)
    df["hardware"] = NEEDLE_HARDWARE
    df["benchmark_suite"] = NEEDLE_SUITE
    df["backend"] = df["backend"].replace({"transformers": "huggingface"})
    df = df.rename(columns={"prompt_tokens": "input_tokens"})
    for col in [
        "model",
        "task",
        "input_tokens",
        "generated_tokens",
        "latency_s",
        "tokens_per_second",
        "rss_before_mb",
        "rss_after_mb",
        "reference",
        "output",
    ]:
        if col not in df.columns:
            df[col] = None
    return df


def build_needle_quality(runs: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for row in runs.to_dict("records"):
        reference = normalize(row.get("reference"))
        output = normalize(row.get("output"))
        has_reference = bool(reference)
        rows.append(
            {
                "experiment": row["experiment"],
                "technique": row["technique"],
                "backend": row["backend"],
                "model": row.get("model"),
                "prompt_id": row["prompt_id"],
                "length_bucket": row["length_bucket"],
                "task": row.get("task"),
                "run": row["run"],
                "exact_match_to_baseline": None,
                "char_similarity_to_baseline": None,
                "rougeL_to_baseline": None,
                "reference_exact_match": float(output == reference) if has_reference else None,
                "reference_contains_match": float(reference in output) if has_reference else None,
                "hardware": NEEDLE_HARDWARE,
                "benchmark_suite": NEEDLE_SUITE,
            }
        )
    return pd.DataFrame(rows)


def build_needle_summary(runs: pd.DataFrame, quality: pd.DataFrame) -> pd.DataFrame:
    df = runs.copy()
    for col in ["latency_s", "tokens_per_second", "input_tokens", "generated_tokens"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    q = quality.copy()
    for col in ["reference_exact_match", "reference_contains_match"]:
        q[col] = pd.to_numeric(q[col], errors="coerce")

    group_cols = ["experiment", "technique", "backend", "length_bucket", "task"]
    perf = (
        df.groupby(group_cols, dropna=False)
        .agg(
            latency_p50=("latency_s", "median"),
            latency_p99=("latency_s", lambda values: values.quantile(0.99)),
            tokens_per_second=("tokens_per_second", "mean"),
            generated_tokens=("generated_tokens", "mean"),
            input_tokens=("input_tokens", "mean"),
            n_runs=("latency_s", "count"),
        )
        .reset_index()
    )
    qual = (
        q.groupby(group_cols, dropna=False)
        .agg(
            reference_exact_match=("reference_exact_match", "mean"),
            reference_contains_match=("reference_contains_match", "mean"),
        )
        .reset_index()
    )
    out = perf.merge(qual, on=group_cols, how="left")
    out["rougeL_to_baseline"] = None
    out["char_similarity_to_baseline"] = None
    out["exact_match_to_baseline"] = None
    out["hardware"] = NEEDLE_HARDWARE
    out["benchmark_suite"] = NEEDLE_SUITE
    return out


def infer_dataset(prompt_id: str) -> str:
    if prompt_id.startswith("cnn_"):
        return "cnn_dailymail"
    if prompt_id.startswith("govreport_"):
        return "govreport"
    if prompt_id.startswith("gsm8k_"):
        return "gsm8k"
    if prompt_id.startswith("needle_"):
        return "needle_in_a_haystack"
    return ""


def read_jsonl_text(text: str) -> list[dict]:
    return [json.loads(line) for line in text.splitlines() if line.strip()]


def load_original_prompts() -> pd.DataFrame:
    """Load the exact 60-prompt set used by the original branch workbook."""
    try:
        text = subprocess.check_output(
            ["git", "show", f"{SOURCE_BRANCH}:prompts/eval_prompts.jsonl"],
            text=True,
        )
        rows = read_jsonl_text(text)
    except subprocess.CalledProcessError:
        rows = read_jsonl_text(Path("prompts/eval_prompts.jsonl").read_text(encoding="utf-8"))
        rows = [row for row in rows if not str(row.get("id", "")).startswith("needle_")]

    df = pd.DataFrame(rows)
    if "dataset" not in df.columns:
        df["dataset"] = df["id"].map(infer_dataset)
    df["hardware"] = ORIGINAL_HARDWARE
    df["benchmark_suite"] = ORIGINAL_SUITE
    df["source"] = f"{SOURCE_BRANCH}:prompts/eval_prompts.jsonl"
    return df


def load_needle_prompts() -> pd.DataFrame:
    df = pd.read_json("prompts/needle_eval_prompts.jsonl", lines=True)
    df["hardware"] = NEEDLE_HARDWARE
    df["benchmark_suite"] = NEEDLE_SUITE
    df["source"] = "prompts/needle_eval_prompts.jsonl"
    return df


def load_all_prompts() -> pd.DataFrame:
    columns = ["id", "length_bucket", "task", "dataset", "reference", "prompt", "hardware", "benchmark_suite", "source"]
    original = load_original_prompts().reindex(columns=columns)
    needle = load_needle_prompts().reindex(columns=columns)
    return concat_aligned([original, needle], columns)


def add_hardware_columns(df: pd.DataFrame, hardware: str, suite: str) -> pd.DataFrame:
    df = df.copy()
    if "hardware" not in df.columns:
        df["hardware"] = hardware
    else:
        df["hardware"] = df["hardware"].fillna(hardware)
    if "benchmark_suite" not in df.columns:
        df["benchmark_suite"] = suite
    else:
        df["benchmark_suite"] = df["benchmark_suite"].fillna(suite)
    return df


def drop_existing_needle_rows(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    mask = pd.Series(False, index=df.index)
    if "benchmark_suite" in df.columns:
        mask |= df["benchmark_suite"].eq(NEEDLE_SUITE)
    if "experiment" in df.columns:
        mask |= df["experiment"].astype(str).str.startswith("needle_")
    return df.loc[~mask].copy()


def union_columns(left: pd.DataFrame, right: pd.DataFrame) -> list[str]:
    cols = list(left.columns)
    for col in right.columns:
        if col not in cols:
            cols.append(col)
    return cols


def concat_aligned(frames: list[pd.DataFrame], columns: list[str]) -> pd.DataFrame:
    usable = [frame.reindex(columns=columns) for frame in frames if len(frame)]
    if not usable:
        return pd.DataFrame(columns=columns)
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="The behavior of DataFrame concatenation with empty or all-NA entries is deprecated")
        return pd.concat(usable, ignore_index=True)


def build_notes() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "topic": "Workbook scope",
                "note": "This is the single canonical benchmark workbook: original summarization/reasoning rows plus Needle all-variants rows.",
            },
            {
                "topic": "Hardware",
                "note": f"Original workbook rows use {ORIGINAL_HARDWARE}; Needle rows use {NEEDLE_HARDWARE}. Do not compare raw latency across hardware without this caveat.",
            },
            {
                "topic": "Needle baseline/KV",
                "note": "Needle HF baseline and HF KV rows use Hugging Face Transformers; long rows use the same CPU fallback policy and are comparable to each other.",
            },
            {
                "topic": "Needle quantized",
                "note": "Needle quantized rows are Ollama GGUF Q4_K_M fallback rows, not HF Metal INT4 rows.",
            },
            {
                "topic": "KV interpretation",
                "note": "Needle is long-input, short-output retrieval. It mostly measures prompt prefill, so KV recency-window is a workload-fit negative result rather than a universal KV-cache failure.",
            },
            {
                "topic": "Memory metrics",
                "note": "RSS columns are process-memory proxies, not true model VRAM measurements.",
            },
            {
                "topic": "Prompt dataset",
                "note": "The Prompts sheet embeds the exact 60 original prompts from experiments/baseline-kvcache plus the 30 Needle prompts.",
            },
        ]
    )


def build_implemented_tasks() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "source": SOURCE_BRANCH,
                "area": "Prompt datasets",
                "implemented": "60-prompt original benchmark set",
                "details": "10 prompts per bucket for summarization and reasoning: CNN/DailyMail short/medium, GovReport long, GSM8K short/medium/long.",
            },
            {
                "source": SOURCE_BRANCH,
                "area": "Ollama baseline",
                "implemented": "baseline_gemma4_e4b",
                "details": "Gemma4:e4b through Ollama, Q4_K_M GGUF, num_ctx=8192, num_predict=128, runs_per_prompt=3.",
            },
            {
                "source": SOURCE_BRANCH,
                "area": "Ollama KV-cache approximation",
                "implemented": "kvcache_limited_512",
                "details": "Gemma4:e4b through Ollama with num_ctx=512, treated as a practical context-window/KV pressure experiment.",
            },
            {
                "source": SOURCE_BRANCH,
                "area": "HF SinkCache/sliding-window",
                "implemented": "hf_baseline and hf_sinkcache",
                "details": "Qwen2.5 HF baseline and sliding-window cache comparison, max input 1024, max new tokens 32.",
            },
            {
                "source": SOURCE_BRANCH,
                "area": "Quality metrics",
                "implemented": "ROUGE-L, char similarity, exact match to baseline",
                "details": "Quality is output-vs-baseline for the original benchmark rows, not reference-answer scoring.",
            },
            {
                "source": "main",
                "area": "Needle all variants",
                "implemented": "needle_hf_baseline, needle_ollama_quantized, needle_hf_kv_window",
                "details": "30 Needle retrieval prompts, reference exact/contains scoring, Apple M3 Pro hardware.",
            },
        ]
    )


def write_workbook(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        tab_colors = {
            "Summary": "1F4E79",
            "Latency p50 (s)": "70AD47",
            "Throughput (tok-s)": "70AD47",
            "ROUGE-L vs baseline": "70AD47",
            "Quality (per run)": "FFC000",
            "All Runs": "FFC000",
            "Prompts": "5B9BD5",
            "Implemented Tasks": "5B9BD5",
            "Hardware & Notes": "7030A0",
        }

        for ws in writer.book.worksheets:
            if ws.title in tab_colors:
                ws.sheet_properties.tabColor = tab_colors[ws.title]
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E79")
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            for col in ws.columns:
                width = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(width + 3, 12), 48)
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if isinstance(cell.value, float):
                        cell.number_format = "0.000"
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default=DEFAULT_INPUT_WORKBOOK)
    parser.add_argument("--out", default=DEFAULT_OUTPUT_WORKBOOK)
    parser.add_argument("--downloads-out", default=DEFAULT_INPUT_WORKBOOK)
    args = parser.parse_args()

    workbook = pd.read_excel(args.input, sheet_name=None)
    for sheet in ["Summary", "Quality (per run)", "All Runs"]:
        workbook[sheet] = add_hardware_columns(
            drop_existing_needle_rows(workbook[sheet]),
            hardware=ORIGINAL_HARDWARE,
            suite=ORIGINAL_SUITE,
        )

    needle_runs = load_needle_runs()
    needle_quality = build_needle_quality(needle_runs)
    needle_summary = build_needle_summary(needle_runs, needle_quality)

    summary_cols = union_columns(workbook["Summary"], needle_summary)
    quality_cols = union_columns(workbook["Quality (per run)"], needle_quality)
    all_runs_cols = union_columns(workbook["All Runs"], needle_runs)

    workbook["Summary"] = concat_aligned([workbook["Summary"], needle_summary], summary_cols)
    workbook["Quality (per run)"] = concat_aligned([workbook["Quality (per run)"], needle_quality], quality_cols)
    workbook["All Runs"] = concat_aligned([workbook["All Runs"], needle_runs], all_runs_cols)

    combined_summary = workbook["Summary"]
    workbook["Latency p50 (s)"] = (
        combined_summary.pivot_table(index="experiment", columns="length_bucket", values="latency_p50", aggfunc="median")
        .reset_index()
    )
    workbook["Throughput (tok-s)"] = (
        combined_summary.pivot_table(index="experiment", columns="length_bucket", values="tokens_per_second", aggfunc="median")
        .reset_index()
    )
    rouge = workbook.get("ROUGE-L vs baseline", pd.DataFrame()).copy()
    if len(rouge):
        rouge = drop_existing_needle_rows(rouge)
    needle_rouge = pd.DataFrame(
        {
            "experiment": sorted(needle_summary["experiment"].unique()),
            "long": None,
            "medium": None,
            "short": None,
        }
    )
    workbook["ROUGE-L vs baseline"] = concat_aligned([rouge, needle_rouge], union_columns(rouge, needle_rouge))
    workbook["Prompts"] = load_all_prompts()
    workbook["Implemented Tasks"] = build_implemented_tasks()
    hardware = pd.DataFrame(
        [
            {
                "hardware": ORIGINAL_HARDWARE,
                "benchmark_suite": ORIGINAL_SUITE,
                "applies_to": "original workbook rows",
            },
            {
                "hardware": NEEDLE_HARDWARE,
                "benchmark_suite": NEEDLE_SUITE,
                "applies_to": "Needle all-variants rows in Summary/All Runs and Prompts sheet",
            },
        ]
    )
    workbook["Hardware & Notes"] = pd.concat(
        [
            hardware.assign(topic="Hardware", note=hardware["applies_to"]),
            build_notes().assign(hardware=None, benchmark_suite=None, applies_to=None),
        ],
        ignore_index=True,
    )[["topic", "hardware", "benchmark_suite", "applies_to", "note"]]

    ordered = [
        "Summary",
        "Latency p50 (s)",
        "Throughput (tok-s)",
        "ROUGE-L vs baseline",
        "Quality (per run)",
        "All Runs",
        "Prompts",
        "Implemented Tasks",
        "Hardware & Notes",
    ]
    sheets = {name: workbook[name] for name in ordered}
    write_workbook(Path(args.out), sheets)
    if args.downloads_out:
        write_workbook(Path(args.downloads_out), sheets)
    print(f"Wrote {args.out}")
    print(f"Wrote {args.downloads_out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
